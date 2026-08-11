"""Text extraction utilities for document types supported by /ingest/doc.

Every extractor here takes a ``max_chars`` ceiling.  The upload size limit
bounds the *compressed* input only; office formats are zip containers, so the
decompressed output has no natural bound.  Each extractor therefore stops
reading as soon as it has produced enough text, rather than building the whole
result and checking its length afterwards.
"""
import logging
import os
from typing import Any

logger = logging.getLogger(__name__)

# Kept in sync with app.services.doc_extraction.MAX_EXTRACTED_CHARS by the
# caller, which always passes the value explicitly.
DEFAULT_MAX_CHARS = 5_000_000

TRUNCATION_MARKER = "\n\n[truncated: document exceeded the extraction size limit]"


class _CharBudget:
    """Accumulate text fragments and stop once the output ceiling is reached."""

    def __init__(self, max_chars: int) -> None:
        self._max_chars = max(1, max_chars)
        self._parts: list[str] = []
        self._length = 0
        self.truncated = False

    @property
    def exhausted(self) -> bool:
        return self._length >= self._max_chars

    @property
    def empty(self) -> bool:
        return not self._parts

    def add_block(self, fragment: str) -> bool:
        """Append a fragment as a blank-line separated block."""

        return self.add(fragment if self.empty else f"\n\n{fragment}")

    def add(self, fragment: str) -> bool:
        """Append a fragment; return False once no more text is wanted."""

        if self.exhausted:
            self.truncated = True
            return False
        remaining = self._max_chars - self._length
        if len(fragment) > remaining:
            self._parts.append(fragment[:remaining])
            self._length = self._max_chars
            self.truncated = True
            return False
        self._parts.append(fragment)
        self._length += len(fragment)
        return True

    def text(self, separator: str = "") -> str:
        joined = separator.join(self._parts)
        return joined + TRUNCATION_MARKER if self.truncated else joined


def extract_document(
    path: str,
    filename: str,
    *,
    max_chars: int = DEFAULT_MAX_CHARS,
) -> tuple[str, dict[str, Any]]:
    """Dispatch to the extractor for the file's extension, output-bounded."""

    ext = os.path.splitext(filename.lower())[1]

    if ext == ".pdf":
        return extract_pdf(path, max_chars=max_chars)
    if ext == ".docx":
        return extract_docx(path, max_chars=max_chars)
    if ext == ".xlsx":
        return extract_xlsx(path, max_chars=max_chars)
    if ext in (".md", ".txt"):
        return extract_text_file(path, max_chars=max_chars)

    raise ValueError(f"Unsupported file extension: {ext}")


def extract_pdf(path: str, *, max_chars: int = DEFAULT_MAX_CHARS) -> tuple[str, dict]:
    """Extract text from a .pdf file, stopping at the output ceiling."""
    import fitz  # pymupdf

    doc = fitz.open(path)
    try:
        metadata: dict[str, Any] = {
            "page_count": len(doc),
            "file_size_bytes": os.path.getsize(path),
        }
        info = doc.metadata or {}
        if info.get("title"):
            metadata["doc_title"] = info["title"]
        if info.get("author"):
            metadata["doc_author"] = info["author"]
            metadata["author"] = info["author"]

        budget = _CharBudget(max_chars)
        for index in range(len(doc)):
            page_text = doc[index].get_text()
            if not page_text.strip():
                continue
            if not budget.add_block(page_text):
                break
    finally:
        doc.close()

    text = budget.text()
    metadata["word_count"] = len(text.split())
    if budget.truncated:
        metadata["extraction_truncated"] = True
    return text, metadata


def extract_docx(path: str, *, max_chars: int = DEFAULT_MAX_CHARS) -> tuple[str, dict]:
    """Extract text from a .docx file.

    Returns (text, metadata) where metadata contains core_properties if available.
    """
    from docx import Document  # python-docx

    doc = Document(path)
    budget = _CharBudget(max_chars)
    for paragraph in doc.paragraphs:
        if not paragraph.text.strip():
            continue
        if not budget.add_block(paragraph.text):
            break
    text = budget.text()

    metadata: dict = {}
    props = doc.core_properties
    if props.title:
        metadata["doc_title"] = props.title
    if props.author:
        metadata["doc_author"] = props.author
    if budget.truncated:
        metadata["extraction_truncated"] = True

    return text, metadata


def extract_xlsx(path: str, *, max_chars: int = DEFAULT_MAX_CHARS) -> tuple[str, dict]:
    """Extract text from all sheets of an .xlsx file, filtering empty rows.

    Returns (text, metadata) where text is a concatenation of all sheets as
    tab-separated rows, and metadata contains sheet names.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    budget = _CharBudget(max_chars)
    sheet_names: list[str] = []

    try:
        for sheet in wb.worksheets:
            sheet_names.append(sheet.title)
            if budget.exhausted:
                continue
            wrote_header = False
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                # Skip rows that are entirely empty
                if not any(c.strip() for c in cells):
                    continue
                if not wrote_header:
                    if not budget.add_block(f"[Sheet: {sheet.title}]"):
                        break
                    wrote_header = True
                if not budget.add("\n" + "\t".join(cells)):
                    break
            if budget.truncated:
                break
    finally:
        wb.close()

    metadata = {"sheets": sheet_names}
    if budget.truncated:
        metadata["extraction_truncated"] = True
    return budget.text(), metadata


def extract_text_file(path: str, *, max_chars: int = DEFAULT_MAX_CHARS) -> tuple[str, dict]:
    """Read a plain-text file (.md, .txt) with UTF-8/latin-1 fallback."""
    for encoding in ("utf-8", "latin-1"):
        try:
            with open(path, encoding=encoding) as fh:
                # Read one character past the ceiling so truncation is detectable.
                text = fh.read(max_chars + 1)
            if len(text) > max_chars:
                return text[:max_chars] + TRUNCATION_MARKER, {"extraction_truncated": True}
            return text, {}
        except UnicodeDecodeError:
            continue
    logger.warning("extract_text_file: could not decode %s with any encoding", path)
    return "", {}
